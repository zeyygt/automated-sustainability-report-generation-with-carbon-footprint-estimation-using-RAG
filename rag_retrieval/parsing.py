"""Layout-aware parsers with fast PDF extraction and OCR fallback hooks."""

from __future__ import annotations

import csv
import datetime as dt
import itertools
import math
import re
from collections.abc import Iterable
from pathlib import Path
from statistics import median
from typing import Protocol

from .config import ParserConfig
from .models import BoundingBox, DocumentInput, ElementType, ParsedDocument, ParsedElement


SPREADSHEET_EXTENSIONS = {".csv", ".xlsx", ".xlsm"}


class DocumentParser(Protocol):
    def parse(self, document: DocumentInput) -> ParsedDocument:
        """Parse one input document into ordered layout elements."""


class OCRFallback(Protocol):
    def extract_text(self, path: Path, max_pages: int) -> str:
        """Return OCR text for documents that have no embedded text."""


class NoopOCRFallback:
    def extract_text(self, path: Path, max_pages: int) -> str:
        return ""


class FallbackTextParser:
    """Fast parser for text-like files and fallback OCR text."""

    def parse(self, document: DocumentInput) -> ParsedDocument:
        text = document.path.read_text(encoding="utf-8", errors="ignore")
        return self.parse_text(document, text)

    def parse_text(self, document: DocumentInput, text: str) -> ParsedDocument:
        elements: list[ParsedElement] = []
        section_path: tuple[str, ...] = ()
        paragraph_buffer: list[str] = []
        counter = itertools.count(1)

        def flush_paragraph() -> None:
            paragraph = _normalize_whitespace("\n".join(paragraph_buffer))
            paragraph_buffer.clear()
            if not paragraph:
                return
            index = next(counter)
            elements.append(
                ParsedElement(
                    element_id=f"{document.doc_id}:text:{index}",
                    doc_id=document.doc_id,
                    page=1,
                    element_type=ElementType.PARAGRAPH,
                    text=paragraph,
                    section_path=section_path,
                )
            )

        for line in text.splitlines():
            if _looks_like_markdown_heading(line):
                flush_paragraph()
                heading = line.lstrip("#").strip()
                section_path = (heading,)
                index = next(counter)
                elements.append(
                    ParsedElement(
                        element_id=f"{document.doc_id}:text:{index}",
                        doc_id=document.doc_id,
                        page=1,
                        element_type=ElementType.HEADING,
                        text=heading,
                        section_path=section_path,
                    )
                )
                continue
            if not line.strip():
                flush_paragraph()
                continue
            paragraph_buffer.append(line)

        flush_paragraph()
        return ParsedDocument(document.doc_id, document.filename, elements, document.metadata)


class SpreadsheetParser:
    """Parser for runtime spreadsheet uploads.

    Workbooks are converted to the same table elements used by PDF parsing so
    chunking, embedding, indexing, and retrieval stay format-agnostic.
    """

    def parse(self, document: DocumentInput) -> ParsedDocument:
        extension = document.path.suffix.lower()
        if extension == ".csv":
            return self._parse_csv(document)
        if extension in {".xlsx", ".xlsm"}:
            return self._parse_xlsx(document)
        raise ValueError(f"Unsupported spreadsheet input: {document.filename}")

    def _parse_xlsx(self, document: DocumentInput) -> ParsedDocument:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for Excel parsing. Install with: pip install openpyxl") from exc

        elements: list[ParsedElement] = []
        counter = itertools.count(1)
        workbook = load_workbook(document.path, read_only=True, data_only=True)
        try:
            for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
                sheet_path = (sheet.title,)
                elements.append(
                    ParsedElement(
                        element_id=f"{document.doc_id}:sheet:{sheet_index}:heading:{next(counter)}",
                        doc_id=document.doc_id,
                        page=sheet_index,
                        element_type=ElementType.HEADING,
                        text=sheet.title,
                        section_path=sheet_path,
                        metadata={"sheet_name": sheet.title, "source_type": "spreadsheet"},
                    )
                )
                rows = [
                    (row_index, [_format_cell(value) for value in row])
                    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1)
                ]
                elements.extend(
                    self._rows_to_table_elements(
                        document=document,
                        rows=rows,
                        sheet_name=sheet.title,
                        sheet_index=sheet_index,
                        section_path=sheet_path,
                        counter=counter,
                    )
                )
        finally:
            workbook.close()

        metadata = dict(document.metadata)
        metadata["parser"] = "spreadsheet"
        return ParsedDocument(document.doc_id, document.filename, elements, metadata)

    def _parse_csv(self, document: DocumentInput) -> ParsedDocument:
        with document.path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = [
                (row_index, [_normalize_whitespace(cell) for cell in row])
                for row_index, row in enumerate(csv.reader(handle), start=1)
            ]

        sheet_name = document.path.stem
        section_path = (sheet_name,)
        counter = itertools.count(1)
        elements = [
            ParsedElement(
                element_id=f"{document.doc_id}:csv:heading:{next(counter)}",
                doc_id=document.doc_id,
                page=1,
                element_type=ElementType.HEADING,
                text=sheet_name,
                section_path=section_path,
                metadata={"sheet_name": sheet_name, "source_type": "spreadsheet"},
            )
        ]
        elements.extend(
            self._rows_to_table_elements(
                document=document,
                rows=rows,
                sheet_name=sheet_name,
                sheet_index=1,
                section_path=section_path,
                counter=counter,
            )
        )
        metadata = dict(document.metadata)
        metadata["parser"] = "spreadsheet"
        return ParsedDocument(document.doc_id, document.filename, elements, metadata)

    def _rows_to_table_elements(
        self,
        document: DocumentInput,
        rows: list[tuple[int, list[str]]],
        sheet_name: str,
        sheet_index: int,
        section_path: tuple[str, ...],
        counter,
    ) -> list[ParsedElement]:
        elements: list[ParsedElement] = []
        for table_number, segment in enumerate(_non_empty_row_segments(rows), start=1):
            trimmed = _trim_empty_columns(segment)
            if not trimmed:
                continue
            header_offset = _find_header_offset(trimmed)
            title_rows = trimmed[:header_offset]
            header_row_index, raw_headers = trimmed[header_offset]
            headers = _normalize_headers(raw_headers)
            body_rows = tuple(
                tuple(_pad(row_values, len(headers))[: len(headers)])
                for _, row_values in trimmed[header_offset + 1 :]
                if any(row_values)
            )
            if not headers or not body_rows:
                continue

            first_data_row = trimmed[header_offset + 1][0] if len(trimmed) > header_offset + 1 else header_row_index
            last_row = trimmed[-1][0]
            min_col, max_col = _segment_column_bounds(segment)
            cell_range = f"{_column_label(min_col + 1)}{header_row_index}:{_column_label(max_col + 1)}{last_row}"
            table_text = _table_to_text(headers, body_rows)
            title = " ".join(_normalize_whitespace(" ".join(row_values)) for _, row_values in title_rows if any(row_values))
            metadata = {
                "headers": headers,
                "rows": body_rows,
                "sheet_name": sheet_name,
                "cell_range": cell_range,
                "table_number": table_number,
                "source_type": "spreadsheet",
                "first_data_row": first_data_row,
                "last_data_row": last_row,
            }
            if title:
                metadata["title"] = title

            elements.append(
                ParsedElement(
                    element_id=f"{document.doc_id}:sheet:{sheet_index}:table:{table_number}:{next(counter)}",
                    doc_id=document.doc_id,
                    page=sheet_index,
                    element_type=ElementType.TABLE,
                    text=table_text,
                    section_path=section_path,
                    metadata=metadata,
                )
            )
        return elements


class RuntimeDocumentParser:
    """Dispatches runtime uploads to the right parser while preserving one API."""

    def __init__(self, config: ParserConfig | None = None, ocr: OCRFallback | None = None):
        self.config = config or ParserConfig()
        self.pdf_parser = PyMuPDFLayoutParser(self.config, ocr)
        self.spreadsheet_parser = SpreadsheetParser()

    def parse(self, document: DocumentInput) -> ParsedDocument:
        if document.path.suffix.lower() in SPREADSHEET_EXTENSIONS:
            return self.spreadsheet_parser.parse(document)
        return self.pdf_parser.parse(document)


class PyMuPDFLayoutParser:
    """PDF parser that preserves headings, paragraphs, reading order, and tables.

    The implementation intentionally uses PyMuPDF's native text and table APIs
    before considering OCR. This keeps the common case fast for born-digital PDFs.
    """

    def __init__(self, config: ParserConfig | None = None, ocr: OCRFallback | None = None):
        self.config = config or ParserConfig()
        self.ocr = ocr or NoopOCRFallback()
        self.text_parser = FallbackTextParser()

    def parse(self, document: DocumentInput) -> ParsedDocument:
        extension = document.path.suffix.lower()
        if extension in {".txt", ".md", ".markdown"}:
            return self.text_parser.parse(document)
        if extension in {".png", ".jpg", ".jpeg", ".webp"}:
            return self._parse_image(document)
        if extension != ".pdf":
            raise ValueError(f"Unsupported parser input: {document.filename}")

        try:
            import fitz  # type: ignore
        except ImportError:
            if self.config.enable_ocr_fallback:
                return self._parse_ocr_text(document)
            raise RuntimeError("PyMuPDF is required for PDF parsing. Install with: pip install pymupdf")

        elements: list[ParsedElement] = []
        with fitz.open(document.path) as pdf:
            element_counter = itertools.count(1)
            for page_index, page in enumerate(pdf, start=1):
                page_elements = self._parse_pdf_page(page, document.doc_id, page_index, element_counter)
                elements.extend(page_elements)
        elements = _assign_sections(elements)

        char_count = sum(len(element.text) for element in elements)
        if char_count < self.config.min_extracted_chars_for_no_ocr and self.config.enable_ocr_fallback:
            ocr_document = self._parse_ocr_text(document)
            if sum(len(element.text) for element in ocr_document.elements) > char_count:
                return ocr_document

        metadata = dict(document.metadata)
        metadata["parser"] = "pymupdf-layout"
        return ParsedDocument(document.doc_id, document.filename, elements, metadata)

    def _parse_pdf_page(self, page, doc_id: str, page_number: int, counter) -> list[ParsedElement]:
        table_elements, table_bboxes = self._extract_tables(page, doc_id, page_number, counter)
        text_elements = self._extract_text_blocks(page, doc_id, page_number, counter, table_bboxes)
        combined = [*table_elements, *text_elements]
        combined.sort(key=lambda element: (element.bbox.y0 if element.bbox else 0.0, element.bbox.x0 if element.bbox else 0.0))
        return combined

    def _extract_tables(self, page, doc_id: str, page_number: int, counter) -> tuple[list[ParsedElement], list[tuple[float, float, float, float]]]:
        if not self.config.detect_tables or not hasattr(page, "find_tables"):
            return [], []

        table_elements: list[ParsedElement] = []
        table_bboxes: list[tuple[float, float, float, float]] = []
        try:
            finder = page.find_tables(strategy=self.config.table_strategy)
        except TypeError:
            finder = page.find_tables()
        except Exception:
            return [], []

        for table_number, table in enumerate(getattr(finder, "tables", []), start=1):
            rows = _clean_table_rows(table.extract())
            if not rows:
                continue
            headers = tuple(str(cell) for cell in rows[0])
            body_rows = tuple(tuple(str(cell) for cell in row) for row in rows[1:])
            bbox = tuple(float(value) for value in table.bbox)
            table_bboxes.append(bbox)
            table_text = _table_to_text(headers, body_rows)
            table_elements.append(
                ParsedElement(
                    element_id=f"{doc_id}:p{page_number}:table:{table_number}:{next(counter)}",
                    doc_id=doc_id,
                    page=page_number,
                    element_type=ElementType.TABLE,
                    text=table_text,
                    bbox=BoundingBox(page_number, *bbox),
                    metadata={"headers": headers, "rows": body_rows, "table_number": table_number},
                )
            )
        return table_elements, table_bboxes

    def _extract_text_blocks(self, page, doc_id: str, page_number: int, counter, table_bboxes) -> list[ParsedElement]:
        try:
            data = page.get_text("dict", sort=True)
        except TypeError:
            data = page.get_text("dict")
        blocks = [block for block in data.get("blocks", []) if block.get("type") == 0]
        font_sizes = [
            span.get("size", 0)
            for block in blocks
            for line in block.get("lines", [])
            for span in line.get("spans", [])
            if span.get("text", "").strip()
        ]
        body_size = median(font_sizes) if font_sizes else 10.0

        elements: list[ParsedElement] = []
        for block in blocks:
            bbox = tuple(float(value) for value in block.get("bbox", (0, 0, 0, 0)))
            if _bbox_inside_any(bbox, table_bboxes):
                continue
            text, max_size, boldish = _block_text_and_style(block)
            text = _normalize_whitespace(text)
            if not text:
                continue
            element_type = ElementType.HEADING if _looks_like_heading(text, max_size, body_size, boldish) else ElementType.PARAGRAPH
            elements.append(
                ParsedElement(
                    element_id=f"{doc_id}:p{page_number}:text:{next(counter)}",
                    doc_id=doc_id,
                    page=page_number,
                    element_type=element_type,
                    text=text,
                    bbox=BoundingBox(page_number, *bbox),
                    metadata={"max_font_size": max_size, "body_font_size": body_size},
                )
            )
        return elements

    def _parse_image(self, document: DocumentInput) -> ParsedDocument:
        if not self.config.enable_ocr_fallback:
            return ParsedDocument(document.doc_id, document.filename, [], document.metadata)
        return self._parse_ocr_text(document)

    def _parse_ocr_text(self, document: DocumentInput) -> ParsedDocument:
        text = self.ocr.extract_text(document.path, self.config.max_ocr_pages)
        if not text:
            return ParsedDocument(
                document.doc_id,
                document.filename,
                [],
                {**document.metadata, "parser": "ocr-unavailable"},
            )
        synthetic = DocumentInput(document.doc_id, document.path, document.filename, "text/plain", document.metadata)
        parsed = self.text_parser.parse_text(synthetic, text)
        return ParsedDocument(parsed.doc_id, parsed.filename, parsed.elements, {**parsed.metadata, "parser": "ocr-fallback"})


def _split_paragraphs(text: str) -> list[str]:
    return [_normalize_whitespace(part) for part in re.split(r"\n\s*\n+", text) if _normalize_whitespace(part)]


def _looks_like_markdown_heading(text: str) -> bool:
    return bool(re.match(r"^\s{0,3}#{1,6}\s+\S+", text))


def _block_text_and_style(block: dict) -> tuple[str, float, bool]:
    lines: list[str] = []
    max_size = 0.0
    boldish = False
    for line in block.get("lines", []):
        fragments = []
        for span in line.get("spans", []):
            text = span.get("text", "")
            if text.strip():
                fragments.append(text)
                max_size = max(max_size, float(span.get("size", 0.0)))
                font_name = str(span.get("font", "")).lower()
                boldish = boldish or "bold" in font_name or bool(span.get("flags", 0) & 16)
        if fragments:
            lines.append(" ".join(fragments))
    return "\n".join(lines), max_size, boldish


def _looks_like_heading(text: str, font_size: float, body_size: float, boldish: bool) -> bool:
    if _looks_like_prose_intro(text):
        return False
    if len(text) > 140 or text.endswith("."):
        return False
    if re.match(r"^\d+(\.\d+)*\s+\S+", text):
        return True
    if font_size >= body_size + 2.0:
        return True
    if boldish and len(text.split()) <= 12:
        return True
    uppercase_letters = sum(1 for char in text if char.isupper())
    letters = sum(1 for char in text if char.isalpha())
    return letters > 4 and uppercase_letters / max(letters, 1) > 0.75 and len(text.split()) <= 12


def _looks_like_prose_intro(text: str) -> bool:
    lowered = text.lower().strip()
    prose_prefixes = (
        "the following table ",
        "the following figure ",
        "the following chart ",
        "this table ",
        "this figure ",
        "this chart ",
        "the table ",
        "the figure ",
        "the chart ",
    )
    return lowered.startswith(prose_prefixes)


def _assign_sections(elements: Iterable[ParsedElement]) -> list[ParsedElement]:
    section_path: tuple[str, ...] = ()
    assigned: list[ParsedElement] = []
    for element in elements:
        if element.element_type == ElementType.HEADING:
            level = _heading_level(element.text)
            section_parts = list(section_path[: max(level - 1, 0)])
            section_parts.append(element.text)
            section_path = tuple(section_parts)
            assigned.append(_replace_section_path(element, section_path))
        else:
            assigned.append(_replace_section_path(element, section_path))
    return assigned


def _heading_level(text: str) -> int:
    match = re.match(r"^(\d+(?:\.\d+)*)\s+", text)
    if match:
        return min(match.group(1).count(".") + 1, 6)
    return 1


def _replace_section_path(element: ParsedElement, section_path: tuple[str, ...]) -> ParsedElement:
    return ParsedElement(
        element_id=element.element_id,
        doc_id=element.doc_id,
        page=element.page,
        element_type=element.element_type,
        text=element.text,
        section_path=section_path,
        bbox=element.bbox,
        metadata=element.metadata,
    )


def _bbox_inside_any(bbox: tuple[float, float, float, float], boxes: list[tuple[float, float, float, float]]) -> bool:
    x0, y0, x1, y1 = bbox
    cx = (x0 + x1) / 2.0
    cy = (y0 + y1) / 2.0
    return any(tx0 <= cx <= tx1 and ty0 <= cy <= ty1 for tx0, ty0, tx1, ty1 in boxes)


def _clean_table_rows(rows) -> list[list[str]]:
    cleaned: list[list[str]] = []
    for row in rows or []:
        values = [_normalize_whitespace("" if cell is None else str(cell)) for cell in row]
        if any(values):
            cleaned.append(values)
    return cleaned


def _table_to_text(headers: tuple[str, ...], rows: tuple[tuple[str, ...], ...]) -> str:
    lines = [" | ".join(headers)]
    lines.append(" | ".join("---" for _ in headers))
    for row in rows:
        padded = [*row, *([""] * max(0, len(headers) - len(row)))]
        lines.append(" | ".join(padded[: len(headers)]))
    return "\n".join(lines)


def _format_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        if value.is_integer():
            return str(int(value))
        formatted = f"{value:,.6f}".rstrip("0").rstrip(".")
        return formatted
    return _normalize_whitespace(str(value))


def _non_empty_row_segments(rows: list[tuple[int, list[str]]]) -> list[list[tuple[int, list[str]]]]:
    segments: list[list[tuple[int, list[str]]]] = []
    current: list[tuple[int, list[str]]] = []
    for row_index, row_values in rows:
        if any(value.strip() for value in row_values):
            current.append((row_index, row_values))
            continue
        if current:
            segments.append(current)
            current = []
    if current:
        segments.append(current)
    return segments


def _trim_empty_columns(segment: list[tuple[int, list[str]]]) -> list[tuple[int, list[str]]]:
    if not segment:
        return []
    min_col, max_col = _segment_column_bounds(segment)
    return [
        (row_index, row_values[min_col : max_col + 1])
        for row_index, row_values in segment
    ]


def _segment_column_bounds(segment: list[tuple[int, list[str]]]) -> tuple[int, int]:
    non_empty_cols = [
        index
        for _, row_values in segment
        for index, value in enumerate(row_values)
        if value.strip()
    ]
    if not non_empty_cols:
        return 0, 0
    return min(non_empty_cols), max(non_empty_cols)


def _find_header_offset(rows: list[tuple[int, list[str]]]) -> int:
    for offset, (_, row_values) in enumerate(rows):
        non_empty_count = sum(1 for value in row_values if value.strip())
        if non_empty_count >= 2:
            return offset
    return 0


def _normalize_headers(raw_headers: list[str]) -> tuple[str, ...]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, raw_header in enumerate(raw_headers):
        header = _normalize_whitespace(raw_header) or f"Column {_column_label(index + 1)}"
        count = seen.get(header, 0) + 1
        seen[header] = count
        if count > 1:
            header = f"{header} {count}"
        headers.append(header)
    return tuple(headers)


def _pad(values: list[str], length: int) -> list[str]:
    return [*values, *([""] * max(0, length - len(values)))]


def _column_label(number: int) -> str:
    label = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        label = chr(65 + remainder) + label
    return label or "A"


def _normalize_whitespace(text: str) -> str:
    text = re.sub(r"[ \t]+", " ", text.replace("\u00a0", " "))
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()
